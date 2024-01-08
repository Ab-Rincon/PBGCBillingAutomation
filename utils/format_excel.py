import openpyxl
import logging
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles import Alignment
from config.settings import template_sheets, summary_loc, problem_loc, detail_loc, overbilled_loc  # noqa
from openpyxl.utils import get_column_letter


def combine_borders(original_border, new_border):
    """
    Combine two borders without overwriting.
    """
    left_style = new_border.left.style if new_border.left else 'none'
    right_style = new_border.right.style if new_border.right else 'none'
    top_style = new_border.top.style if new_border.top else 'none'
    bottom_style = new_border.bottom.style if new_border.bottom else 'none'

    combined = Border(
        left=new_border.left if left_style != 'none' else original_border.left,
        right=new_border.right if right_style != 'none' else original_border.right,
        top=new_border.top if top_style != 'none' else original_border.top,
        bottom=new_border.bottom if bottom_style != 'none' else original_border.bottom,
        diagonal=original_border.diagonal,
        diagonal_direction=original_border.diagonal_direction,
        outline=original_border.outline,
        vertical=original_border.vertical,
        horizontal=original_border.horizontal
    )
    return combined


def insert_page_break(worksheet):
    # Assuming you want the page break after column C, for instance
    col_num = 6
    last_row = worksheet.max_row  # Get the last row with data

    worksheet.print_area = f"A1:F{last_row}"
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.fitToWidth = 1

    worksheet.col_breaks.append(openpyxl.worksheet.pagebreak.Break(id=col_num))
    worksheet.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=last_row))

    # Save the workbook with the new page break
    logging.info("Inserted page breaks in summary sheet")


def add_summary_name_block(worksheet, locations):
    # Loop through the rows of the worksheet
    for row in range(locations['Summary']["row"] + 1, worksheet.max_row + 1):  # Assuming you start checking from the second row
        # Check if the value in column B of the current row is different from the previous row
        if worksheet[f"B{row}"].value != worksheet[f"B{row-1}"].value:
            # If it is, apply the border to columns B through E for the current row
            for col in ['B', 'C', 'D', 'E']:
                # Get the current borders of the cell
                current_border = worksheet[f"{col}{row}"].border
                # Update the top border while maintaining other borders
                worksheet[f"{col}{row}"].border = Border(
                    top=Side(style='thin'),
                    left=current_border.left,
                    right=current_border.right,
                    bottom=current_border.bottom
                )
    logging.info("Inserted name dividers in summary sheet")


def format_sheet(workbook, workbook_filename, worksheet, template_sheet, locations, dataframes):
    # Create a white fill for the current worksheet
    fill_white_worksheet(worksheet)

    # Create borders
    create_border_and_align_worsheet(template_sheet, worksheet, locations, dataframes)

    # Auto size columns
    autosize_columns_worksheet(worksheet, template_sheet)

    # Save workbook
    workbook.save(workbook_filename)
    logging.info(f'Formatted {worksheet}')


def format_all_code_sheets(workbook_filename, dataframes, invoice_sheet_name):
    # Load workbook
    workbook = openpyxl.load_workbook(workbook_filename)

    for template_sheet in template_sheets:
        if template_sheet == "Mismatch":
            continue

        # Determine the current worksheet
        try:
            worksheet = workbook[f'{invoice_sheet_name}_{template_sheet}']
        except Exception as e:
            logging.debug(f'Unable to find the worksheet for template: {template_sheet}. Exception: {e}')
            continue

        # Determine location of items in worksheets
        locations = globals()[f'{template_sheet.lower()}_loc']

        # Apply general formatting
        format_sheet(workbook, workbook_filename, worksheet, template_sheet, locations, dataframes)

        # Wrap comment in col Y
        # if template_sheet == 'Detail':
        #    wrap_text_col(worksheet, 'X', 'X2')
        #    wrap_text_col(worksheet, 'Y', 'Y2')

        # Do special formatting for the summary worksheet
        if template_sheet == 'Summary':

            # Make borders for summary sheet
            add_summary_name_block(worksheet, locations)

            # Insert page break to export pdf
            insert_page_break(worksheet)

            # Make column E wrap text in summary sheet
            wrap_text_col(worksheet, 'E', 'E9')

            # Set page break header
            set_page_break_header(worksheet)

            # Created banded rows for readability
            create_summary_banded_rows(worksheet, locations)

            # Top align rows
            top_align_summary_rows(worksheet, locations)

        workbook.save(workbook_filename)
    logging.info("Completed formatting for charge code sheets\n")


def fill_white_worksheet(worksheet):
    # Define white fill
    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    # Assuming we want to fill up to column 'Z' (for example) which is the 26th column
    # If you want to fill until the last column you expect data could be in, set this higher
    max_column_to_fill = worksheet.max_column + 10

    # To find the maximum touched column (by style or value), we loop through cells
    # We set this to a minimum of the existing max column in case there's styled but empty cells beyond it
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1):
        for cell in row:
            if cell.value or cell.has_style:  # Check if cell has value or style
                max_column_to_fill = max(max_column_to_fill, cell.column)

    # Now iterate over the range of rows and columns to apply the fill
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row + 50, min_col=1, max_col=max_column_to_fill):
        for cell in row:
            cell.fill = white_fill

    logging.debug(f"Applied white fill to the worksheet up to column: {get_column_letter(max_column_to_fill)}")


def create_border_and_align_worsheet(template_sheet, worksheet, locations, dataframes):

    for key in locations:
        # Apply no borders if there's no data
        if dataframes[key].empty:
            continue

        # Determine starting location
        if template_sheet == "Summary":
            start_row = locations[key]["row"] - 1
        else:
            start_row = locations[key]["row"] - 2
        start_col = locations[key]["col"]

        # Center titles for data
        for cell in worksheet[start_row + 1]:
            if template_sheet == "Summary":
                continue
            cell.alignment = Alignment(horizontal="center")

        # Find the end column for the data
        end_col = start_col
        while worksheet.cell(row=start_row + 1, column=end_col).value and end_col <= worksheet.max_column:
            end_col += 1
        end_col -= 1

        # Find the end row for the data
        end_row = start_row
        while worksheet.cell(row=end_row, column=start_col).value and end_row <= worksheet.max_row:
            end_row += 1
        end_row -= 1

        # Define thick borders
        thick_border_top = Border(top=Side(style='medium'))
        thick_border_bottom = Border(bottom=Side(style='medium'))
        thick_border_left = Border(left=Side(style='medium'))
        thick_border_right = Border(right=Side(style='medium'))

        # Apply top and bottom borders
        for col in range(start_col, end_col + 1):
            current_cell = worksheet.cell(row=start_row, column=col)
            current_cell.border = combine_borders(current_cell.border, thick_border_top)

            current_cell = worksheet.cell(row=end_row, column=col)
            current_cell.border = combine_borders(current_cell.border, thick_border_bottom)

        # Apply left and right borders
        for row in range(start_row, end_row):
            current_cell = worksheet.cell(row=row, column=start_col)
            current_cell.border = combine_borders(current_cell.border, thick_border_left)

            current_cell = worksheet.cell(row=row, column=end_col)
            current_cell.border = combine_borders(current_cell.border, thick_border_right)

        # Special handling for bottom left and right corners
        bottom_left_cell = worksheet.cell(row=end_row, column=start_col)
        bottom_left_cell.border = combine_borders(bottom_left_cell.border, thick_border_left)
        bottom_left_cell.border = combine_borders(bottom_left_cell.border, thick_border_bottom)

        bottom_right_cell = worksheet.cell(row=end_row, column=end_col)
        bottom_right_cell.border = combine_borders(bottom_right_cell.border, thick_border_right)
        bottom_right_cell.border = combine_borders(bottom_right_cell.border, thick_border_bottom)

        # Save the changes
        logging.debug(f"Applied Borders and aligned titled in sheet: {worksheet}")


def autosize_columns_worksheet(worksheet, template_sheet):
    # Loop through each column in worksheet using iter_cols
    for col in worksheet.iter_cols():
        max_length = 0
        column = col  # No need to list comprehension here; 'col' is already a tuple of cells.
        for cell in column:
            # Skip rows based on template_sheet condition.
            if template_sheet == "Summary" and cell.row < 9:
                continue
            try:
                # Try to find the length of the cell value
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))  # Convert cell value to string before measuring length.
            except Exception as e:
                logging.debug(f"Error processing cell: {cell} with error: {e}")
                pass  # In case of any error, continue without breaking the loop.

        # Set a minimum width and a maximum width for the column.
        adjusted_width = max((max_length + 2), 10)  # Set a minimum width to 10
        adjusted_width = min(adjusted_width, 68)  # Set a maximum width to 68 to prevent very wide columns.

        # Set the column width using column_dimensions and get_column_letter
        if column:  # Check if the column tuple is not empty.
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Log the debug message after all columns have been adjusted.
    logging.debug("Adjusted column sizes for template_sheet: {}".format(template_sheet))


def wrap_text_col(worksheet, column, align_cell=None):
    # Set wrap text style for each cell in column E
    for cell in worksheet[column]:  # This will iterate over all cells in column E
        cell.alignment = Alignment(wrap_text=True)
    if align_cell is not None:
        cell = worksheet[align_cell]
        cell.alignment = Alignment(horizontal="center")


def set_page_break_header(worksheet):
    # Set rows 2 through 9 to repeat at top
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1  # Adjusts to fit width on one page if needed
    worksheet.print_title_rows = '2:9'  # Rows to repeat at the top of every printed page


def create_summary_banded_rows(worksheet, locations, band_size=1):
    # Get locations
    start_row = locations['Summary']['row']
    start_col = locations['Summary']['col']
    end_col = 5  # Col E
    end_row = worksheet.max_row

    # Define the fill styles for banded rows
    odd_fill = PatternFill(start_color='00CCCCCC', end_color='00CCCCCC', fill_type='solid')
    even_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    # Apply the fill styles
    for row in range(start_row, end_row + 1, band_size * 2):  # Start from start_row, go up by double the band size
        for band_row in range(row, min(row + band_size, end_row + 1)):  # Apply the odd_fill for band_size rows
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=band_row, column=col)
                if worksheet.cell(row=band_row, column=2).value not in (None, ""):  # Check if the cell in column B is not empty
                    cell.fill = odd_fill

        for band_row in range(row + band_size, min(row + band_size * 2, end_row + 1)):  # Apply the even_fill for the next band_size rows
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=band_row, column=col)
                if worksheet.cell(row=band_row, column=2).value not in (None, ""):  # Check if the cell in column B is not empty
                    cell.fill = even_fill


def top_align_summary_rows(worksheet, locations):
    # Get locations of cells
    start_row = locations['Summary']['row']
    start_col = locations['Summary']['col']
    end_col = 4  # Col D
    end_row = worksheet.max_row

    # Top align only column A
    top_alignment = Alignment(vertical='top')
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.alignment = top_alignment
