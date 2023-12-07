import pandas as pd, logging, os
import openpyxl
import xlwings as xw
import utils.clean_data as cd
from config.settings import template_sheets, summary_loc, problem_loc, detail_loc, overbilled_loc, blacklist_charge_codes

def set_logger():
    """
    Configure the logger for logging messages.
    """
    # Get the root logger
    logger = logging.getLogger()
    while logger.hasHandlers():
        logger.removeHandler(logger.handlers[0])

    logger.setLevel(logging.INFO)

    # Set the log format
    formatter = logging.Formatter("%(asctime)s %(levelname)s: %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    # Set up console logging
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    logger.addHandler(console_handler)

    # Set up file logging
    file_handler = logging.FileHandler('logfile.log', mode="w")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

def read_excel_data(filename, sheet_name) -> pd.DataFrame:
    logging.info(f"The current sheet is: {sheet_name}")
    # Read the first column to determine where the data starts
    df_full = pd.read_excel(f'input/{filename}', sheet_name=sheet_name, usecols=[0], engine='openpyxl')
    
    # Find the row index where the column 'A' has the value 'Name', this row is likely before the headers
    try: name_row_idx = df_full[df_full.iloc[:,0] == 'Name'].index[0]
    except Exception:
        logging.exception('A name header was not located. Please ensure the headers are correct.')
        raise IndexError(error_text)     
    
    # The header is expected to be right after the 'Name' row
    header_row_idx = name_row_idx + 1
    if header_row_idx > 10: 
        error_text = f'Please insure that the proper headers are include in the input file. It appears a little high at row {name_row_idx}'
        logging.warn(error_text)

    # Find the last non-empty cell in the first column to determine how many rows of data there are
    non_empty_rows = df_full.iloc[:,0].dropna().index
    last_data_row_idx = non_empty_rows[-1] if non_empty_rows.size > 0 else header_row_idx

    # Calculate number of rows to read
    nrows = last_data_row_idx - header_row_idx + 1

    # Now read the required range using the correct header row
    df = pd.read_excel(
        f'input/{filename}',
        sheet_name=sheet_name,
        usecols=[0, 2, 3, 4],  # Adjust as needed to read the correct columns
        skiprows=list(range(header_row_idx)),  # Skip all rows up to the header row
        nrows=nrows,
        header=0,  # Now the first row of data read will be treated as the header
        engine='openpyxl'
    )

    # Convert the column to datetime format (if not already)
    try: df['Date'] = pd.to_datetime(df['Date'])
    except Exception: 
        error_text = 'There was an error in converting the date column to date format. Please review the input file to ensure every row in the date column is in fact a date.'
        logging.exception(error_text)
        raise TypeError(error_text)

    # Reformat the datetime to MM/DD/YYYY format
    df['Date'] = df['Date'].dt.strftime('%m/%d/%Y')

    # Add charge code to dataframe to track charge codes
    df['Charge Code'] = sheet_name

    logging.info(f'There are {len(df.index)} total comments')

    return df

def copy_and_rename_excel(filename, invoice_sheet_names):
    # Filenames
    destination_filename = filename.split(" Invoice")[0]
    source_path = 'config/sheet_template.xlsx'
    destination_path = f'output/{destination_filename} Organized Invoice.xlsx'

    # If destination file does not exist, create it by copying the source file
    if not os.path.exists(destination_path):
        with open(source_path, 'rb') as source_file:
            with open(destination_path, 'wb') as dest_file:
                dest_file.write(source_file.read())
    
    # Start Excel in the background
    app = xw.App(visible=False)
    
    # Open both workbooks using the App instance
    source_wb = app.books.open(source_path)
    target_wb = app.books.open(destination_path)

    for invoice_sheet_name in invoice_sheet_names:
        for sheet_name in template_sheets:
            if sheet_name == "Mismatch": continue

            # Access the source sheet
            source_sheet = source_wb.sheets[sheet_name]

            # Copy the source sheet to the target workbook if it's not the first itteration
            if invoice_sheet_name != invoice_sheet_names[0]: source_sheet.api.Copy(Before=target_wb.sheets[0].api)

            # Rename the copied sheet in the target workbook
            target_wb.sheets[sheet_name].name = f'{invoice_sheet_name}_{sheet_name}'

            # Save and close the target workbook
            target_wb.save()
    target_wb.close()

    # Close the source workbook without saving changes
    source_wb.close()
    
    # Quit the app instance to close Excel
    app.quit()

    logging.info(f"Appended sheets from {source_path} to {destination_path}.")
    return destination_path

def paste_all_to_excel(dataframes: dict, excel_file, invoice_sheet_name):
    # Load the existing workbook
    workbook = openpyxl.load_workbook(excel_file)
    
    all_df_empty = True
    for key, df in dataframes.items():
        if key == "Summary" or key == "Mismatch" or df.empty: continue

        #Check if the dataframe is empty
        if key != 'Acceptable': all_df_empty = False

        # Define problem and detail sheet names
        problem_sheet_name = f'{invoice_sheet_name}_Problem'
        detail_sheet_name = f'{invoice_sheet_name}_Detail'

        # Load or create the problem sheet by name
        problem_sheet = workbook[problem_sheet_name]
        detail_sheet = workbook[detail_sheet_name]

        # Paste problem DataFrame headers and data for 'Name' and 'Date' columns
        problem_headers = df[['Name', 'Date']].columns
        for c, header in enumerate(problem_headers, start=problem_loc[key]['col']):
            problem_sheet.cell(row=problem_loc[key]['row']-1, column=c).value = header

        for r, row_data in enumerate(df[['Name', 'Date']].values, start=problem_loc[key]['row']):
            for c, value in enumerate(row_data, start=problem_loc[key]['col']):
                problem_sheet.cell(row=r, column=c).value = value

        # Paste detail DataFrame headers and data
        detail_headers = df.columns
        for c, header in enumerate(detail_headers, start=detail_loc[key]['col']):
            detail_sheet.cell(row=detail_loc[key]['row']-1, column=c).value = header

        for r, row_data in enumerate(df.values, start=detail_loc[key]['row']):
            for c, value in enumerate(row_data, start=detail_loc[key]['col']):
                detail_sheet.cell(row=r, column=c).value = value
        logging.debug(f"Pasted sheet {key}")
    
    # If all dataframes are empty, delete the sheets
    if all_df_empty:
        for sheet_name in [f'{invoice_sheet_name}_Problem', f'{invoice_sheet_name}_Detail']:
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
                workbook.save(excel_file)

    # Paste summary DataFrame headers and data
    summary_sheet_name = f'{invoice_sheet_name}_Summary'
    summary_sheet = workbook[summary_sheet_name]
    key = "Summary"
    df = dataframes[key].drop_duplicates(subset=["Name", "Date", "Formatted Time Comments"])
    if df.empty:
        del workbook[summary_sheet_name]
        workbook.save(excel_file)
        return all_df_empty

    for r, row_data in enumerate(df[['Name', 'Date', 'Total Hours Worked','Formatted Time Comments']].values, start=summary_loc[key]['row']):
        for c, value in enumerate(row_data, start=summary_loc[key]['col']):
            summary_sheet.cell(row=r, column=c).value = value
    
    #Insert Month in summary
    month = excel_file.split()[1]
    year = excel_file.split()[2]
    month_text = f'Month: {month} {year}'
    #contract_number = f'Contract Number: {invoice_sheet_name}'

    #Insert contract number
    #summary_sheet["B4"] = contract_number

    # Insert the value into the specified cell
    summary_sheet["B6"] = month_text

    # Save the workbook
    workbook.save(excel_file)
    logging.info("Pasted problematic data into relevant sheets")
    return all_df_empty

def list_visible_sheets_in_workbook(workbook_path) -> tuple[str, pd.DataFrame]:
    logging.info(f"The current workbook is: {workbook_path}")
    destination_filename = workbook_path.split(" Invoice")[0]
    destination_path = f'output/{destination_filename} Organized Invoice.xlsx'

    # Load the workbook
    workbook = openpyxl.load_workbook(f'input/{workbook_path}')

    # Get the list of visible sheet names
    visible_sheet_names = [sheet.title for sheet in workbook if sheet.sheet_state == 'visible']

    # Filter out the sheet names that are in the blacklist
    visible_sheet_names = [sheet for sheet in visible_sheet_names if sheet not in blacklist_charge_codes]

    # Delete destination workbook if it exists
    try:
        os.remove(destination_path)
        print(f"{destination_path} has been deleted.")
    except FileNotFoundError:
        print(f"{destination_path} not found.")
    
    logging.info(f'List of visible sheets: {visible_sheet_names}')
    return visible_sheet_names

def find_workbook_list():
    import os

    # Specify the directory path
    folder_path = 'input/'

    # Get a list of all files in the directory that do not start with "~"
    file_list = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f)) and not f.startswith('~')]
    return file_list

def create_overbilled_sheet(overbilled_df: pd.DataFrame, worksheet:str):
    key = "Mismatch"
    empty_df = False
    
    # Paste summary DataFrame headers and data
    if overbilled_df.empty:
        del worksheet
        return True
    
    # Add data to overbilled worksheet including headers
    for c, header in enumerate(overbilled_df.columns, start=overbilled_loc[key]['col']):
        worksheet.cell(row=overbilled_loc[key]['row']-1, column=c).value = header

    for r, row_data in enumerate(overbilled_df.values, start=overbilled_loc[key]['row']):
        for c, value in enumerate(row_data, start=overbilled_loc[key]['col']):
            worksheet.cell(row=r, column=c).value = value
    return empty_df